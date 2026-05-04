"""
generate_varied_questions.py — 從測試題庫種子生成多樣化「多輪對話」

功能：
1. 讀取「聯評聊天機器人測試.xlsx」中的 78 道種子問題（按 A~M 分類）
2. 以「類別」為單位，利用 LLM 生成多組完整多輪對話 session（3~5 輪）
3. 同類別內：同主題深入追問 / 子類切換 / 口語+書面混搭
4. 跨類別：模擬家長自然話題轉換（如 B 分數解讀 → E 在家訓練）
5. 對弱勢類別（I, K, M, A, D）加倍生成
6. 匯出為 auto_query_bot.py 相容的 SCENARIOS 格式

用法：
    python rl_pipeline/scripts/generate_varied_questions.py
    # 產出: rl_pipeline/scripts/generated_scenarios.py
    #       rl_pipeline/scripts/generated_scenarios.json
"""

import os
import sys
import json
import time
import re
import random
import openpyxl
from collections import Counter, defaultdict
from dotenv import load_dotenv

load_dotenv()
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '../..')))

from openai import OpenAI

# --- 設定 ---
EXCEL_PATH = r"C:\Users\88696\Desktop\edu_sys\聯評聊天機器人測試.xlsx"
OUTPUT_PATH = os.path.join(os.path.dirname(__file__), "generated_scenarios.py")
OUTPUT_JSON_PATH = os.path.join(os.path.dirname(__file__), "generated_scenarios.json")

MODEL = "gpt-4.1-mini"

# 每個類別要生成幾組「同類別多輪對話」
SESSIONS_PER_CATEGORY = 4
BOOST_CATEGORIES = ["I", "K", "M", "A", "D"]  # 弱勢類別
BOOST_EXTRA_SESSIONS = 3  # 弱勢類別額外幾組

# 跨類別對話組合數
CROSS_CATEGORY_SESSIONS = 2  # 每對生成幾組

CATEGORY_NAMES = {
    "A": "報告總覽與閱讀順序",
    "B": "PT分數/量表/百分位解讀",
    "C": "PT臨床觀察行為描述",
    "D": "能力剖面（強弱項分析）",
    "E": "在家PT訓練建議",
    "F": "融入日常粗大動作練習",
    "G": "是否需要PT/早療評估",
    "H": "轉介在地資源/機構查詢",
    "I": "報告分享/隱私與安全",
    "J": "與學校合作",
    "K": "補助/福利申請",
    "L": "追蹤再評估",
    "M": "家長情緒支持與壓力調適",
}

# 跨類別組合（模擬常見話題轉換）
CROSS_CATEGORY_PAIRS = [
    ("A", "B"), ("B", "E"), ("B", "N"), ("C", "D"), ("C", "E"),
    ("D", "E"), ("E", "F"), ("E", "M"), ("G", "H"), ("H", "K"),
    ("I", "J"), ("J", "K"), ("L", "B"), ("L", "N"), ("F", "M"),
    ("A", "G"), ("C", "G"), ("D", "F"), ("M", "G"), ("B", "D"),
]


def read_seed_questions(excel_path: str) -> dict:
    """從 Excel 讀取種子問題，按類別分組"""
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb["Questions"]
    by_category = defaultdict(list)
    for r in range(2, ws.max_row + 1):
        qid = ws.cell(r, 1).value
        cat = ws.cell(r, 2).value
        sub = ws.cell(r, 3).value
        q = ws.cell(r, 4).value
        intent = ws.cell(r, 5).value
        domain = ws.cell(r, 6).value
        if qid and cat and q and isinstance(qid, str) and qid.startswith("Q"):
            cat_letter = cat.strip()[0] if cat else "?"
            seed = {
                "id": qid,
                "category_letter": cat_letter,
                "category": cat,
                "subcategory": sub or "",
                "question": q,
                "intent": intent or "",
                "domain": domain or "",
            }
            by_category[cat_letter].append(seed)
    return dict(by_category)


def _clean_lines(raw: str) -> list:
    """解析 LLM 輸出，移除編號，按空行分組為多組對話"""
    sessions = []
    current = []
    for line in raw.split("\n"):
        line = line.strip()
        line = re.sub(r"^[\d]+[\.\)、:\s]+", "", line).strip()
        # 移除「---」分隔線或「對話N」等標頭
        if re.match(r"^[-=]{3,}$", line) or re.match(r"^(對話|Session|Conversation)\s*\d*", line, re.I):
            if current:
                sessions.append(current)
                current = []
            continue
        if not line:
            if current:
                sessions.append(current)
                current = []
            continue
        current.append(line)
    if current:
        sessions.append(current)
    return sessions


def generate_intra_category_sessions(
    client: OpenAI, cat_letter: str, seeds: list, n_sessions: int
) -> list:
    """為同一類別生成 n_sessions 組完整多輪對話（每組 3~5 輪）"""
    cat_name = CATEGORY_NAMES.get(cat_letter, cat_letter)
    seed_questions = "\n".join(
        f"- [{s['subcategory']}] {s['question']}（意圖：{s['intent']}）"
        for s in seeds
    )

    prompt = f"""你是一位早療評估報告諮詢系統的測試設計師。

以下是「{cat_name}」類別（{cat_letter}）的種子問題：
{seed_questions}

請以這些種子為基礎，生成 {n_sessions} 組完整的「家長多輪對話」。

【規則】
1. 每組對話 3~5 輪（每輪 = 家長問的一句話）
2. 第 1 輪以種子問題的某個改寫版開場
3. 後續輪次要自然地追問或延伸，像真實對話：
   - 追問細節：「那具體是哪方面？」
   - 要求解釋：「什麼意思？」
   - 問做法：「那我在家可以怎麼做？」
   - 表達擔心：「這樣算嚴重嗎？」
   - 子類切換：從分數問到觀察、從觀察問到建議
4. 風格要多變：口語（帶語氣詞）、書面、簡短追問、帶情境、爸爸/阿嬤/老師角度
5. 每組對話之間用空行分隔
6. 不同組的第 1 輪要用不同的種子問題做為起點
7. 不要加編號、不要加角色標籤、不要加說明文字
8. 只輸出家長說的話，不要輸出系統/助手的回覆

請直接輸出 {n_sessions} 組對話："""

    try:
        response = client.chat.completions.create(
            model=MODEL,
            messages=[
                {"role": "system", "content": "你是早療對話測試專家。只輸出家長的提問，每組對話用空行分隔，不加編號或說明。"},
                {"role": "user", "content": prompt},
            ],
            temperature=0.9,
            max_tokens=2000,
        )
        raw = response.choices[0].message.content.strip()
        sessions = _clean_lines(raw)
        # 只保留 2~6 輪的對話
        valid = [s for s in sessions if 2 <= len(s) <= 6]
        return valid[:n_sessions]
    except Exception as e:
        print(f"  [ERROR] 同類別對話生成失敗 ({cat_letter}): {e}")
        return []


def generate_cross_category_sessions(
    client: OpenAI, cat1: str, cat2: str,
    seeds1: list, seeds2: list, n_sessions: int
) -> list:
    """生成跨類別的多輪對話（先問 cat1 再自然轉向 cat2）"""
    name1 = CATEGORY_NAMES.get(cat1, cat1)
    name2 = CATEGORY_NAMES.get(cat2, cat2)

    sample1 = "\n".join(f"- {s['question']}" for s in seeds1[:4])
    sample2 = "\n".join(f"- {s['question']}" for s in seeds2[:4])

    prompt = f"""你是一位早療評估報告諮詢系統的測試設計師。

請設計 {n_sessions} 組跨主題的「家長多輪對話」，模擬家長從「{name1}」自然轉到「{name2}」的情境。

【{name1} ({cat1}) 參考問題】
{sample1}

【{name2} ({cat2}) 參考問題】
{sample2}

【規則】
1. 每組 4~5 輪
2. 前 1~2 輪圍繞 {name1}（{cat1}）
3. 中間自然過渡（如「對了，我還想問...」「那這跟...有關嗎？」「另外想請教...」）
4. 後 2~3 輪圍繞 {name2}（{cat2}）
5. 風格多變：口語、書面、帶情境、簡短追問
6. 每組用空行分隔
7. 不加編號、不加角色標籤、不加說明
8. 只輸出家長的話

請直接輸出 {n_sessions} 組對話："""

    try:
        response = client.chat.completions.create(
            model=MODEL,
            messages=[
                {"role": "system", "content": "你是早療對話測試專家。輸出跨主題多輪對話，每組空行分隔，不加編號或說明。"},
                {"role": "user", "content": prompt},
            ],
            temperature=0.9,
            max_tokens=2000,
        )
        raw = response.choices[0].message.content.strip()
        sessions = _clean_lines(raw)
        valid = [s for s in sessions if 3 <= len(s) <= 6]
        return valid[:n_sessions]
    except Exception as e:
        print(f"  [ERROR] 跨類別對話生成失敗 ({cat1}+{cat2}): {e}")
        return []


def build_all_scenarios(
    by_category: dict,
    intra_sessions: dict,
    cross_sessions: list,
) -> list:
    """將所有生成的對話整合為 SCENARIOS 格式"""
    scenarios = []

    # 1. 同類別多輪對話
    for cat, sessions in sorted(intra_sessions.items()):
        cat_name = CATEGORY_NAMES.get(cat, cat)
        for i, steps in enumerate(sessions):
            scenarios.append({
                "name": f"{cat}-{cat_name}-對話{i+1}",
                "steps": steps,
                "final_feedback": 1,
                "metadata": {"category": cat, "type": "intra_category"},
            })

    # 2. 跨類別多輪對話
    for item in cross_sessions:
        scenarios.append(item)

    # 3. 隨機打亂順序（但保持可重現）
    random.seed(42)
    random.shuffle(scenarios)

    return scenarios


def export_as_python(scenarios: list, output_path: str):
    """匯出為 auto_query_bot.py 可直接 import 的格式"""
    clean = []
    for s in scenarios:
        entry = {"name": s["name"], "steps": s["steps"]}
        if "final_feedback" in s:
            entry["final_feedback"] = s["final_feedback"]
        clean.append(entry)

    with open(output_path, "w", encoding="utf-8") as f:
        f.write('"""\n')
        f.write("自動生成的多樣化多輪測試 SCENARIOS\n")
        f.write(f"生成時間: {time.strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write(f"總情境數: {len(scenarios)}\n")
        total_steps = sum(len(s["steps"]) for s in scenarios)
        f.write(f"總問句數: {total_steps}\n")
        f.write('"""\n\n')
        f.write("GENERATED_SCENARIOS = ")
        f.write(json.dumps(clean, ensure_ascii=False, indent=4))
        f.write("\n")

    print(f"\n[輸出] Python: {output_path}")


def export_as_json(scenarios: list, output_path: str):
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(scenarios, f, ensure_ascii=False, indent=2)
    print(f"[輸出] JSON:   {output_path}")


def main():
    print("=" * 60)
    print(" 多樣化「多輪對話」生成器 v2")
    print("=" * 60)

    # 1. 讀取種子問題
    by_category = read_seed_questions(EXCEL_PATH)
    total_seeds = sum(len(v) for v in by_category.values())
    print(f"\n[讀取] 種子問題: {total_seeds} 題, {len(by_category)} 個類別")
    for cat in sorted(by_category):
        n = len(by_category[cat])
        name = CATEGORY_NAMES.get(cat, "?")
        boost = " ★" if cat in BOOST_CATEGORIES else ""
        print(f"  {cat}: {n} 題 — {name}{boost}")

    # 2. 初始化 LLM
    client = OpenAI()

    # =============================================
    # 第一階段：同類別多輪對話
    # =============================================
    print(f"\n{'='*40}")
    print(" 第一階段：同類別多輪對話")
    print(f"{'='*40}")
    intra_sessions = {}
    total_intra = 0

    for cat in sorted(by_category):
        seeds = by_category[cat]
        n_sessions = SESSIONS_PER_CATEGORY
        if cat in BOOST_CATEGORIES:
            n_sessions += BOOST_EXTRA_SESSIONS

        print(f"\n[{cat}] {CATEGORY_NAMES.get(cat, '?')} — 生成 {n_sessions} 組對話...")

        sessions = generate_intra_category_sessions(client, cat, seeds, n_sessions)
        intra_sessions[cat] = sessions

        for s in sessions:
            print(f"  OK {len(s)} 輪: {s[0][:40]}...")
        total_intra += len(sessions)
        time.sleep(0.5)

    print(f"\n[同類別小計] {total_intra} 組對話")

    # =============================================
    # 第二階段：跨類別多輪對話
    # =============================================
    print(f"\n{'='*40}")
    print(" 第二階段：跨類別多輪對話")
    print(f"{'='*40}")
    cross_sessions = []
    total_cross = 0

    for cat1, cat2 in CROSS_CATEGORY_PAIRS:
        if cat1 not in by_category or cat2 not in by_category:
            continue

        n = CROSS_CATEGORY_SESSIONS
        print(f"\n[{cat1}→{cat2}] {CATEGORY_NAMES.get(cat1,'')} → {CATEGORY_NAMES.get(cat2,'')} — {n} 組...")

        sessions = generate_cross_category_sessions(
            client, cat1, cat2,
            by_category[cat1], by_category[cat2], n
        )

        for i, steps in enumerate(sessions):
            cross_sessions.append({
                "name": f"跨類-{cat1}→{cat2}-對話{i+1}",
                "steps": steps,
                "final_feedback": 1,
                "metadata": {"category": f"{cat1}→{cat2}", "type": "cross_category"},
            })
            print(f"  OK {len(steps)} 輪: {steps[0][:40]}...")
        total_cross += len(sessions)
        time.sleep(0.5)

    print(f"\n[跨類別小計] {total_cross} 組對話")

    # =============================================
    # 第三階段：整合匯出
    # =============================================
    scenarios = build_all_scenarios(by_category, intra_sessions, cross_sessions)

    total_steps = sum(len(s["steps"]) for s in scenarios)
    step_dist = Counter(len(s["steps"]) for s in scenarios)
    type_dist = Counter(s["metadata"]["type"] for s in scenarios)

    print(f"\n{'='*40}")
    print(f" 最終統計")
    print(f"{'='*40}")
    print(f"  總對話 sessions: {len(scenarios)}")
    print(f"  總問句數:        {total_steps}")
    print(f"  輪數分布:        {dict(sorted(step_dist.items()))}")
    print(f"  類型分布:        {dict(type_dist)}")

    # 按類別統計
    cat_dist = Counter()
    for s in scenarios:
        cat_info = s["metadata"]["category"]
        cat_dist[cat_info] += 1
    print(f"  類別分布:")
    for cat, cnt in sorted(cat_dist.items()):
        print(f"    {cat}: {cnt} 組")

    export_as_python(scenarios, OUTPUT_PATH)
    export_as_json(scenarios, OUTPUT_JSON_PATH)

    print(f"\n[完成] 使用方式:")
    print(f"  在 auto_query_bot.py 中:")
    print(f"    from generated_scenarios import GENERATED_SCENARIOS")
    print(f"    SCENARIOS += GENERATED_SCENARIOS")

    # Cleanup
    temp_json = os.path.join(os.path.dirname(__file__), "../../temp_questions.json")
    if os.path.exists(temp_json):
        os.remove(temp_json)


if __name__ == "__main__":
    main()
