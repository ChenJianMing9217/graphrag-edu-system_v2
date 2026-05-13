"""
列出 90 turn 的 domain 預測 + 我（標註者）認為的正確答案，
產出 Excel 給人工標註後用來算 domain 命中率與設計改進方向。

Output: expert_eval/domain_analysis_v1.xlsx
  - Sheet 「Domain 標註表」: 90 列, 每列含 query / system 預測 / 標註欄
  - Sheet 「Domain 對應表」: 11 個 domain 的定義摘要 (給標註者參考)
"""
import json
from pathlib import Path
from collections import Counter

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parent.parent
RAW_DIR = Path(__file__).resolve().parent / "raw_logs"
OUT_PATH = ROOT / "expert_eval" / "domain_analysis_v1.xlsx"
ANCHORS = json.load(open(ROOT / "dialogue_state_module" / "config" / "domain_anchors.json", encoding="utf-8"))
DOMAINS = ANCHORS["domains"]

H_FILL = PatternFill("solid", fgColor="2F5597")
H_FONT = Font(name="Microsoft JhengHei", size=11, bold=True, color="FFFFFF")
SCORE_FILL = PatternFill("solid", fgColor="C00000")
BAD_FILL = PatternFill("solid", fgColor="FFE6E6")
OK_FILL = PatternFill("solid", fgColor="E6FFE6")
BODY = Font(name="Microsoft JhengHei", size=10)
THIN = Side(style="thin", color="999999")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top", horizontal="left")
CENTER = Alignment(wrap_text=True, vertical="center", horizontal="center")


def load_all():
    rows = []
    for fp in sorted(RAW_DIR.glob("CASE_*_G*.json")):
        d = json.load(open(fp, encoding="utf-8"))
        for t in d["turns"]:
            fs = t.get("flow_state") or {}
            rows.append({
                "case_id": d["case_id"],
                "group_key": d["group_key"],
                "focus": d["focus"],
                "role": d["role"],
                "turn_no": t["turn_no"],
                "query": t["query"],
                "task_pred": fs.get("task_pred"),
                "memory_action": fs.get("memory_action"),
                "top_domain": fs.get("top_domain"),
                "top_domain_prob": fs.get("top_domain_prob"),
                "active_domains": fs.get("active_domains") or [],
                "normalized_entropy": fs.get("normalized_entropy"),
                "topic_overlap": fs.get("topic_overlap"),
            })
    return rows


def build_workbook(rows):
    wb = Workbook()
    wb.remove(wb.active)

    # ===== Sheet 1: Domain 標註表 =====
    ws = wb.create_sheet("Domain 標註表", 0)
    headers = [
        ("個案", 9),
        ("題組", 7),
        ("測試重點", 18),
        ("角色", 11),
        ("Turn", 5),
        ("提問", 50),
        ("預測 task", 10),
        ("預測 top_domain", 18),
        ("top_prob", 10),
        ("active_domains", 30),
        ("norm_entropy", 11),
        ("我認為的正確 top_domain", 22),
        ("我認為的可接受 domains\n(逗號分隔)", 28),
        ("判定 (合理/可接受/不合理)", 18),
        ("錯誤類型\n(抽象/sticky/anchor缺/角色)", 22),
        ("備註", 30),
    ]
    for ci, (h, w) in enumerate(headers, start=1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = H_FONT
        c.fill = SCORE_FILL if ci >= 12 else H_FILL
        c.alignment = CENTER
        c.border = BORDER
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 36
    ws.freeze_panes = "G2"

    # Data validations
    dv_judge = DataValidation(
        type="list",
        formula1='"合理,可接受,不合理"',
        allow_blank=True,
    )
    dv_err = DataValidation(
        type="list",
        formula1='"抽象描述,sticky過強,anchor缺失,角色切換,其他"',
        allow_blank=True,
    )
    dv_domain = DataValidation(
        type="list",
        formula1='"' + ",".join(DOMAINS) + '"',
        allow_blank=True,
    )
    ws.add_data_validation(dv_judge)
    ws.add_data_validation(dv_err)
    ws.add_data_validation(dv_domain)

    for ri, r in enumerate(rows, start=2):
        vals = [
            r["case_id"], r["group_key"], r["focus"], r["role"], r["turn_no"],
            r["query"], r["task_pred"], r["top_domain"],
            round(r["top_domain_prob"], 3) if r["top_domain_prob"] else "",
            ", ".join(r["active_domains"]),
            round(r["normalized_entropy"], 3) if r["normalized_entropy"] else "",
            "", "", "", "", "",
        ]
        for ci, v in enumerate(vals, start=1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.font = BODY
            c.alignment = WRAP
            c.border = BORDER
        ws.row_dimensions[ri].height = 72
    last = len(rows) + 1
    dv_judge.add(f"N2:N{last}")
    dv_err.add(f"O2:O{last}")
    dv_domain.add(f"L2:L{last}")

    # ===== Sheet 2: Domain 對應表 =====
    ws2 = wb.create_sheet("Domain 對應表", 1)
    headers2 = [("Domain", 22), ("Anchor 數量", 14), ("Anchor 摘要 (前 3 個)", 80)]
    for ci, (h, w) in enumerate(headers2, start=1):
        c = ws2.cell(row=1, column=ci, value=h)
        c.font = H_FONT
        c.fill = H_FILL
        c.alignment = CENTER
        c.border = BORDER
        ws2.column_dimensions[get_column_letter(ci)].width = w
    for ri, d in enumerate(DOMAINS, start=2):
        anchors = ANCHORS["domain_anchors"].get(d, [])
        ws2.cell(row=ri, column=1, value=d).font = BODY
        ws2.cell(row=ri, column=2, value=len(anchors)).font = BODY
        preview = "\n".join(f"• {a}" for a in anchors[:3])
        ws2.cell(row=ri, column=3, value=preview).font = BODY
        for ci in range(1, 4):
            ws2.cell(row=ri, column=ci).alignment = WRAP
            ws2.cell(row=ri, column=ci).border = BORDER
        ws2.row_dimensions[ri].height = 80

    # ===== Sheet 3: 預測分布統計 =====
    ws3 = wb.create_sheet("預測分布", 2)
    headers3 = [("Domain", 24), ("被預測為 top 的次數", 22), ("出現在 active 的次數", 22)]
    for ci, (h, w) in enumerate(headers3, start=1):
        c = ws3.cell(row=1, column=ci, value=h)
        c.font = H_FONT
        c.fill = H_FILL
        c.alignment = CENTER
        c.border = BORDER
        ws3.column_dimensions[get_column_letter(ci)].width = w

    top_counter = Counter(r["top_domain"] for r in rows)
    active_counter = Counter()
    for r in rows:
        for d in r["active_domains"]:
            active_counter[d] += 1
    for ri, d in enumerate(DOMAINS, start=2):
        ws3.cell(row=ri, column=1, value=d).font = BODY
        ws3.cell(row=ri, column=2, value=top_counter.get(d, 0)).font = BODY
        ws3.cell(row=ri, column=3, value=active_counter.get(d, 0)).font = BODY
        for ci in range(1, 4):
            ws3.cell(row=ri, column=ci).alignment = CENTER
            ws3.cell(row=ri, column=ci).border = BORDER

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_PATH)
    print(f"-> {OUT_PATH}")
    print(f"\n預測分布 (top_domain):")
    for d, n in top_counter.most_common():
        print(f"  {d:<28} {n}")


if __name__ == "__main__":
    rows = load_all()
    print(f"Loaded {len(rows)} turns from {len(set(r['case_id']+r['group_key'] for r in rows))} chats")
    build_workbook(rows)
