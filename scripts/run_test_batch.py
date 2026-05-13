"""
Run multi-turn evaluation test batch against the deployed EDUBOT system.

For each case x each group:
  1. login_with_code  -> set server-side active_child_id
  2. new_chat         -> reset session, new chat_session_id
  3. For each of 6 turns:
       POST /api/chat_stream, accumulate delta, capture done event
  4. Save raw log JSON to scripts/raw_logs/{case_id}_{group_key}.json

Usage:
  python scripts/run_test_batch.py                          # full batch
  python scripts/run_test_batch.py --case=CASE_A            # one case
  python scripts/run_test_batch.py --case=CASE_A --group=G1 # smoke test
"""
import json
import sys
import time
from datetime import datetime
from pathlib import Path

import openpyxl
import requests

ROOT = Path(__file__).resolve().parent.parent
CASES_FILE = Path(__file__).resolve().parent / "cases.json"
TEST_BANK = ROOT / "expert_eval" / "測試題組.xlsx"
OUT_DIR = Path(__file__).resolve().parent / "raw_logs"
OUT_DIR.mkdir(exist_ok=True)

DELAY_BETWEEN_TURNS = 2
TIMEOUT_PER_TURN = 240

GROUP_KEY_MAP = {
    "題組一": "G1",
    "題組二": "G2",
    "題組三": "G3",
    "題組四": "G4",
    "題組五": "G5",
}


def load_test_bank(xlsx_path: Path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    groups = {}
    for r in range(2, ws.max_row + 1):
        group = ws.cell(r, 1).value
        role = ws.cell(r, 2).value
        focus = ws.cell(r, 3).value
        turn_no = ws.cell(r, 4).value
        query = ws.cell(r, 5).value
        if not group or not query:
            continue
        key = GROUP_KEY_MAP.get(group, group)
        if key not in groups:
            groups[key] = {"name": group, "role": role, "focus": focus, "turns": []}
        groups[key]["turns"].append({"turn_no": int(turn_no), "query": query})
    for g in groups.values():
        g["turns"].sort(key=lambda t: t["turn_no"])
    return groups


def login(session: requests.Session, base_url: str, code: str):
    r = session.post(f"{base_url}/api/login_with_code", json={"code": code}, timeout=30)
    r.raise_for_status()
    data = r.json()
    if data.get("status") != "success":
        raise RuntimeError(f"Login failed: {data}")
    return data


def new_chat(session: requests.Session, base_url: str):
    r = session.post(f"{base_url}/api/new_chat", json={}, timeout=30)
    r.raise_for_status()
    return r.json()


def logout(session: requests.Session, base_url: str):
    try:
        session.post(f"{base_url}/api/logout", json={}, timeout=10)
    except Exception:
        pass


def chat_stream(session: requests.Session, base_url: str, query: str):
    t0 = time.time()
    r = session.post(
        f"{base_url}/api/chat_stream",
        json={"message": query, "response_length": "standard"},
        stream=True,
        timeout=TIMEOUT_PER_TURN,
        headers={"Accept": "text/event-stream"},
    )
    r.raise_for_status()

    response_text = ""
    flow_state = None
    retrieval_info = None
    message_id = None
    error = None

    for line in r.iter_lines(decode_unicode=True):
        if not line or not line.startswith("data: "):
            continue
        payload = line[6:]
        try:
            event = json.loads(payload)
        except json.JSONDecodeError:
            continue
        t = event.get("type")
        if t == "delta":
            response_text += event.get("text", "")
        elif t == "done":
            flow_state = event.get("flow_state")
            retrieval_info = event.get("retrieval_info")
            message_id = event.get("message_id")
        elif t == "error":
            error = event.get("error")

    return {
        "response_text": response_text,
        "flow_state": flow_state,
        "retrieval_info": retrieval_info,
        "message_id": message_id,
        "latency_sec": round(time.time() - t0, 2),
        "error": error,
    }


def run_group(case: dict, group_key: str, group_data: dict, base_url: str):
    case_id = case["case_id"]
    code = case["access_code"]
    out_path = OUT_DIR / f"{case_id}_{group_key}.json"

    print(f"\n[{case_id} / {group_key}] {group_data['focus']} ({group_data['role']})")

    session = requests.Session()
    started = datetime.now().isoformat(timespec="seconds")
    try:
        login_data = login(session, base_url, code)
        print(f"  login OK ({login_data.get('child_name')})")
        new_chat(session, base_url)

        results = []
        for turn in group_data["turns"]:
            preview = turn["query"][:40].replace("\n", " ")
            print(f"  T{turn['turn_no']}: {preview}... ", end="", flush=True)
            try:
                res = chat_stream(session, base_url, turn["query"])
                if res.get("error"):
                    print(f"FAIL: {res['error']}")
                else:
                    print(f"OK ({res['latency_sec']}s, {len(res['response_text'])} chars)")
            except Exception as e:
                print(f"EXC: {e}")
                res = {"error": str(e), "latency_sec": None}
            res["turn_no"] = turn["turn_no"]
            res["query"] = turn["query"]
            res["timestamp"] = datetime.now().isoformat(timespec="seconds")
            results.append(res)
            time.sleep(DELAY_BETWEEN_TURNS)

        logout(session, base_url)

        record = {
            "case_id": case_id,
            "access_code": code,
            "group_key": group_key,
            "group_name": group_data["name"],
            "role": group_data["role"],
            "focus": group_data["focus"],
            "started_at": started,
            "finished_at": datetime.now().isoformat(timespec="seconds"),
            "base_url": base_url,
            "turns": results,
        }
        with open(out_path, "w", encoding="utf-8") as f:
            json.dump(record, f, ensure_ascii=False, indent=2)
        print(f"  -> saved {out_path.name}")
        return record
    finally:
        logout(session, base_url)


def main():
    cfg = json.load(open(CASES_FILE, encoding="utf-8"))
    base_url = cfg["base_url"]
    cases = cfg["cases"]
    bank = load_test_bank(TEST_BANK)

    only_case = None
    only_group = None
    for arg in sys.argv[1:]:
        if arg.startswith("--case="):
            only_case = arg.split("=", 1)[1]
        elif arg.startswith("--group="):
            only_group = arg.split("=", 1)[1]

    selected_cases = [c for c in cases if (not only_case or c["case_id"] == only_case)]
    selected_groups = sorted(bank.keys())
    if only_group:
        selected_groups = [g for g in selected_groups if g == only_group]

    print(f"Cases: {[c['case_id'] for c in selected_cases]}")
    print(f"Groups: {selected_groups}")
    print(f"Total chat rooms: {len(selected_cases) * len(selected_groups)}")

    for case in selected_cases:
        for gk in selected_groups:
            try:
                run_group(case, gk, bank[gk], base_url)
            except Exception as e:
                print(f"[ERROR] {case['case_id']}/{gk}: {e}")


if __name__ == "__main__":
    main()
