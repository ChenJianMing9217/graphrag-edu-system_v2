"""
比較 raw_logs_v1/ (anchor 修正前) 與 raw_logs/ (修正後) 的 domain / task / memory 預測。

Output:
  - stdout 摘要
  - expert_eval/domain_compare_v1_v2.xlsx  (4 sheets)
"""
import json
from collections import Counter
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
V1_DIR = Path(__file__).resolve().parent / "raw_logs_v1"
V2_DIR = Path(__file__).resolve().parent / "raw_logs"
OUT = ROOT / "expert_eval" / "domain_compare_v1_v2.xlsx"

# 來自 heuristic_label.py 的關鍵字判定（複用）
DOMAIN_KEYWORDS = {
    "粗大動作": [
        "動作表現", "動作能力", "動作練習", "動作上", "動作發展",
        "走路", "跑", "跳", "平衡", "核心", "穩定", "肌力",
        "上下樓梯", "粗大", "體能課", "大肌肉", "全身",
    ],
    "精細動作": ["精細", "握筆", "寫字", "剪刀", "拼圖", "扣鈕扣",
              "手部", "手指", "雙手協調", "手眼協調", "用湯匙", "用筷子"],
    "感覺統合": ["感覺統合", "感統", "前庭", "本體", "觸覺",
              "感覺尋求", "怕高", "怕聲音"],
    "口腔動作": ["口腔動作", "口肌", "嘴唇", "舌頭", "咀嚼", "吹氣", "流口水"],
    "情緒行為與社會適應功能": [
        "情緒", "哭鬧", "生氣", "崩潰", "脾氣", "焦慮", "擔心",
        "嚴重", "自責", "壓力", "情感", "搶玩具", "社交",
        "不喜歡", "不要做", "說累", "說不", "拒絕", "抗拒",
    ],
    "吞嚥功能": ["吞嚥", "嗆", "嗆到", "進食安全", "含著不吞"],
    "口語理解": ["聽得懂", "聽不懂", "理解指令", "兩步驟", "理解我說",
              "聽指令", "理解語意"],
    "口語表達": ["說話", "表達", "詞彙", "句子", "用講的", "說出來",
              "講不清楚", "回答問題"],
    "說話": ["發音", "構音", "音清楚", "說話不清楚", "可懂度", "口齒"],
    "認知功能": ["注意力", "專注", "記憶", "學習能力", "認知",
              "規則", "步驟", "因果", "概念", "配對", "分類"],
    "整體概況": ["整份報告", "整體", "報告主要", "全部一起", "整體狀況",
              "報告在說什麼", "整理成", "整理重點", "整體評估"],
}


def expected_domains(query):
    return {d for d, kws in DOMAIN_KEYWORDS.items() if any(kw in query for kw in kws)}


def load_dir(dirpath: Path):
    """{(case, group, turn): row}"""
    out = {}
    for fp in sorted(dirpath.glob("CASE_*_G*.json")):
        d = json.load(open(fp, encoding="utf-8"))
        for t in d["turns"]:
            fs = t.get("flow_state") or {}
            key = (d["case_id"], d["group_key"], t["turn_no"])
            out[key] = {
                "case": d["case_id"],
                "group": d["group_key"],
                "focus": d["focus"],
                "role": d["role"],
                "turn": t["turn_no"],
                "query": t["query"],
                "top_domain": fs.get("top_domain"),
                "top_prob": fs.get("top_domain_prob"),
                "active_domains": fs.get("active_domains") or [],
                "task_pred": fs.get("task_pred"),
                "memory_action": fs.get("memory_action"),
            }
    return out


# ------ Helpers ------

H_FILL = PatternFill("solid", fgColor="2F5597")
H_FONT = Font(name="Microsoft JhengHei", size=11, bold=True, color="FFFFFF")
GOOD = PatternFill("solid", fgColor="E6FFE6")
BAD = PatternFill("solid", fgColor="FFE6E6")
NEUTRAL = PatternFill("solid", fgColor="F2F2F2")
BODY = Font(name="Microsoft JhengHei", size=10)
THIN = Side(style="thin", color="999999")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top", horizontal="left")
CENTER = Alignment(wrap_text=True, vertical="center", horizontal="center")


def hdr(ws, row, headers):
    for ci, (h, w) in enumerate(headers, start=1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = H_FONT
        c.fill = H_FILL
        c.alignment = CENTER
        c.border = BORDER
        ws.column_dimensions[get_column_letter(ci)].width = w


def write_row(ws, row, vals, fills=None):
    for ci, v in enumerate(vals, start=1):
        c = ws.cell(row=row, column=ci, value=v if v is not None else "")
        c.font = BODY
        c.alignment = WRAP
        c.border = BORDER
        if fills and ci - 1 < len(fills) and fills[ci - 1]:
            c.fill = fills[ci - 1]


# ------ Main analysis ------

def main():
    v1 = load_dir(V1_DIR)
    v2 = load_dir(V2_DIR)
    common = sorted(v1.keys() & v2.keys())
    print(f"v1: {len(v1)}  v2: {len(v2)}  common: {len(common)}")
    if not common:
        print("ERROR: 沒有共同的 turn (v2 還沒跑完?)")
        return

    # 統計
    flipped = []          # v1 != v2 在 top_domain
    flipped_to_correct = 0
    flipped_to_wrong = 0
    same_correct = 0
    same_wrong = 0

    for k in common:
        r1, r2 = v1[k], v2[k]
        exp = expected_domains(r2["query"])
        if r1["top_domain"] != r2["top_domain"]:
            flipped.append((k, r1, r2, exp))
            if exp:
                in_exp_v1 = r1["top_domain"] in exp
                in_exp_v2 = r2["top_domain"] in exp
                if not in_exp_v1 and in_exp_v2:
                    flipped_to_correct += 1
                elif in_exp_v1 and not in_exp_v2:
                    flipped_to_wrong += 1
        else:
            if exp and r1["top_domain"] in exp:
                same_correct += 1
            elif exp and r1["top_domain"] not in exp:
                same_wrong += 1

    print(f"\n=== Domain 預測變化 ===")
    print(f"  變動: {len(flipped)} / {len(common)} = {len(flipped)/len(common)*100:.1f}%")
    print(f"  flipped to CORRECT (修好): {flipped_to_correct}")
    print(f"  flipped to WRONG   (退化): {flipped_to_wrong}")
    print(f"  same & correct (維持對的): {same_correct}")
    print(f"  same & wrong   (仍是錯的): {same_wrong}")

    # 各 domain 預測分布
    v1_dist = Counter(v1[k]["top_domain"] for k in common)
    v2_dist = Counter(v2[k]["top_domain"] for k in common)
    print(f"\n=== top_domain 分布變化 ===")
    print(f"{'Domain':<28} {'v1':>5} {'v2':>5} {'Δ':>5}")
    for d in sorted(set(v1_dist) | set(v2_dist)):
        a, b = v1_dist.get(d, 0), v2_dist.get(d, 0)
        delta = b - a
        marker = " ↑" if delta > 0 else (" ↓" if delta < 0 else "")
        print(f"  {d:<28} {a:>5} {b:>5} {delta:>+5}{marker}")

    # Memory & Task 變化
    mem_diff = sum(1 for k in common if v1[k]["memory_action"] != v2[k]["memory_action"])
    task_diff = sum(1 for k in common if v1[k]["task_pred"] != v2[k]["task_pred"])
    print(f"\n=== 副作用檢查 ===")
    print(f"  Memory 判斷改變: {mem_diff} / {len(common)}")
    print(f"  Task 預測改變  : {task_diff} / {len(common)}")

    # Build Excel
    wb = Workbook()
    wb.remove(wb.active)

    # Sheet 1: 變動明細
    ws = wb.create_sheet("Domain 變動明細", 0)
    hdr(ws, 1, [
        ("個案", 9), ("題組", 7), ("Turn", 5), ("提問", 50),
        ("v1 top_domain", 18), ("v1 prob", 9),
        ("v2 top_domain", 18), ("v2 prob", 9),
        ("v1→v2", 10), ("關鍵字期望", 25),
    ])
    ws.freeze_panes = "E2"
    row = 2
    for k, r1, r2, exp in flipped:
        v1_correct = r1["top_domain"] in exp if exp else None
        v2_correct = r2["top_domain"] in exp if exp else None
        if exp:
            if not v1_correct and v2_correct:
                tag = "修好 ✓"
                fill = GOOD
            elif v1_correct and not v2_correct:
                tag = "退化 ✗"
                fill = BAD
            else:
                tag = "改變但不確定"
                fill = NEUTRAL
        else:
            tag = "無關鍵字"
            fill = NEUTRAL
        fills = [None] * 8 + [fill, fill]
        write_row(ws, row,
                  [r2["case"], r2["group"], r2["turn"], r2["query"],
                   r1["top_domain"], round(r1["top_prob"] or 0, 3),
                   r2["top_domain"], round(r2["top_prob"] or 0, 3),
                   tag, ", ".join(sorted(exp)) if exp else ""],
                  fills=fills)
        ws.row_dimensions[row].height = 60
        row += 1

    # Sheet 2: 分布對比
    ws2 = wb.create_sheet("分布對比", 1)
    hdr(ws2, 1, [("Domain", 28), ("v1 預測為 top", 14), ("v2 預測為 top", 14), ("變化", 10)])
    row = 2
    for d in sorted(set(v1_dist) | set(v2_dist)):
        a, b = v1_dist.get(d, 0), v2_dist.get(d, 0)
        delta = b - a
        write_row(ws2, row, [d, a, b, delta])
        row += 1

    # Sheet 3: 總體統計
    ws3 = wb.create_sheet("統計摘要", 2)
    hdr(ws3, 1, [("指標", 30), ("數值", 15)])
    summaries = [
        ("總 turn 數", len(common)),
        ("Domain 變動數", len(flipped)),
        ("Domain 變動率", f"{len(flipped)/len(common)*100:.1f}%"),
        ("flipped to CORRECT (修好)", flipped_to_correct),
        ("flipped to WRONG (退化)", flipped_to_wrong),
        ("same & correct (維持對)", same_correct),
        ("same & wrong (仍錯)", same_wrong),
        ("Memory 改變數", mem_diff),
        ("Task 預測改變數", task_diff),
    ]
    for ri, (k, v) in enumerate(summaries, start=2):
        write_row(ws3, ri, [k, v])

    # Sheet 4: Memory/Task 變化（如果有）
    if mem_diff or task_diff:
        ws4 = wb.create_sheet("Memory & Task 變化", 3)
        hdr(ws4, 1, [
            ("個案", 9), ("題組", 7), ("Turn", 5), ("提問", 40),
            ("v1 memory", 12), ("v2 memory", 12),
            ("v1 task", 10), ("v2 task", 10),
        ])
        row = 2
        for k in common:
            r1, r2 = v1[k], v2[k]
            if r1["memory_action"] != r2["memory_action"] or r1["task_pred"] != r2["task_pred"]:
                fills = [None] * 4 + [
                    BAD if r1["memory_action"] != r2["memory_action"] else None,
                    BAD if r1["memory_action"] != r2["memory_action"] else None,
                    BAD if r1["task_pred"] != r2["task_pred"] else None,
                    BAD if r1["task_pred"] != r2["task_pred"] else None,
                ]
                write_row(ws4, row,
                          [r2["case"], r2["group"], r2["turn"], r2["query"],
                           r1["memory_action"], r2["memory_action"],
                           r1["task_pred"], r2["task_pred"]],
                          fills=fills)
                row += 1

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"\n-> {OUT}")


if __name__ == "__main__":
    main()
