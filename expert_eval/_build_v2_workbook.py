"""
重建專家評分工作簿 v2

輸出：expert_eval/初次評分_v2.xlsx

工作表：
  1. 單輪題庫     — 48 既有 + 25 補題（含 F/N/C/E 補強 + 邊界 case + SQL 實測）
  2. 多輪 Scenarios — 8 個多輪對話腳本（共 ~30 turns），測 Memory Agent
  3. 評分說明     — 9 維評估標準

評分維度（9 維）：
  - 資料正確性  / 連貫性 / 可行性 / 安全性 / 完整性 / 整體
  - 引用適當性  (新)
  - 長度適當性  (新)
  - 多輪一致性  (新, 僅多輪 scenario 評)
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

ROOT = Path(__file__).resolve().parent
OUT = ROOT / "初次評分_v2.xlsx"

# ============ 樣式 ============
HEADER_FILL = PatternFill("solid", fgColor="2F5C7E")  # 深藍底
HEADER_FONT = Font(name="Microsoft JhengHei", size=11, bold=True, color="FFFFFF")
CELL_FONT = Font(name="Microsoft JhengHei", size=10)
ASPECT_FILL = PatternFill("solid", fgColor="FFF4E5")
SCENARIO_FILL = PatternFill("solid", fgColor="E8F4FF")
THIN = Side(border_style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="center", horizontal="left")
CENTER = Alignment(horizontal="center", vertical="center")

# ============ 既有 48 題 ============
EXISTING = [
    (1, 1, "看完聯評報告我很亂，先看哪三個重點？", "報告解讀"),
    (2, 2, "如果只能先做一件事，最建議先做什麼？", "決策規劃"),
    (3, 1, "在家看起來還好，為什麼報告說有挑戰？", "報告解讀"),
    (4, 2, "我應該怎麼記錄，治療師才看得懂？", "日常追蹤"),
    (5, 3, "可以給我一個一週觀察清單嗎？", "日常追蹤"),
    (6, 1, "孩子已在早療，為什麼報告還寫有落差？", "介入調整"),
    (7, 2, "這樣要加課還是換方向？", "介入調整"),
    (8, 3, "怎麼判斷是孩子累了還是課程不合適？", "介入調整"),
    (9, 4, "下次回診前我該準備哪些資料？", "醫療流程"),
    (10, 1, "第一次要做聯評，通常先掛哪一科？", "醫療流程"),
    (11, 2, "要帶哪些文件才不會白跑？", "醫療流程"),
    (12, 3, "如果現場孩子不合作，結果會不會失真？", "醫療流程"),
    (13, 1, "孩子抗拒練習，一做就哭，怎麼辦？", "家庭訓練"),
    (14, 2, "可以把訓練改成遊戲嗎？", "家庭訓練"),
    (15, 3, "每天做多久比較不會太累？", "家庭訓練"),
    (16, 4, "幼兒園老師可以怎麼配合？", "家校合作"),
    (17, 5, "怎麼知道這個月有沒有進步？", "日常追蹤"),
    (18, 1, "孩子最近睡不好、很容易崩潰，會影響早療嗎？", "身心狀態"),
    (19, 2, "要先看身心科還是先跟治療師討論？", "醫療流程"),
    (20, 3, "等待門診期間，在家先做什麼比較安全？", "安全與風險"),
    (21, 4, "哪些警訊需要立即就醫？", "安全與風險"),
    (22, 1, "幼小銜接前，一定要再做心理評估嗎？", "學校銜接"),
    (23, 2, "如果沒通過資源班，還有什麼支持？", "學校銜接"),
    (24, 3, "普通班先試讀，我要追蹤哪些指標？", "學校銜接"),
    (25, 1, "不同醫院報告結論不一樣，我該信哪個？", "報告解讀"),
    (26, 2, "可以帶著舊報告去做第二意見嗎？", "醫療流程"),
    (27, 3, "跨院資料常不被採認，事前怎麼確認？", "醫療流程"),
    (28, 4, "我該用什麼格式整理資料最有效？", "日常追蹤"),
    (29, 1, "我被說「家長沒做好」，我很自責怎麼辦？", "家長支持"),
    (30, 2, "怎麼把情緒拉回來，先做好下一步？", "家長支持"),
    (31, 1, "聯評結果在臨界值，這代表什麼？", "報告解讀"),
    (32, 2, "這種情況要先衝密集課程嗎？", "介入調整"),
    (33, 3, "家裡人意見不一樣，怎麼達成共識？", "家長支持"),
    (34, 4, "學校端我該先談哪些支持？", "家校合作"),
    (35, 5, "三個月內我該追什麼量化指標？", "日常追蹤"),
    (36, 6, "若沒進步，下一步決策流程是什麼？", "決策規劃"),
    (37, 1, "報告寫「臨界」到底是接近正常還是接近遲緩？我該怎麼理解風險？", "報告解讀"),
    (38, 2, "同一份報告裡分數和文字結論看起來不一致時，應該以哪個為準？", "報告解讀"),
    (39, 3, "報告沒有寫清楚優先順序，我要怎麼判斷先練哪一項？", "報告解讀"),
    (40, 4, "不同醫院的報告結論不一樣，我該如何整合成同一個訓練方向？", "報告解讀"),
    (41, 5, "報告建議很多，我家時間有限，怎麼排出可執行的前兩項？", "報告解讀"),
    (42, 1, "要找早療機構時，家長應該先看哪些重點（專長、頻率、距離、等候）？", "機構選擇"),
    (43, 2, "如果大醫院要排很久，先去診所做早療會有什麼利弊？", "機構選擇"),
    (44, 3, "不同機構課程名稱很多，我怎麼判斷是否真的符合孩子需求？", "機構選擇"),
    (45, 1, "補助申請最常缺哪些文件？我可以一次備齊什麼？", "補助申請"),
    (46, 2, "沒有身心障礙證明時，還有哪些資源或補助可以先申請？", "補助申請"),
    (47, 3, "補助被退件時，我應該先找哪個窗口確認原因與補件方式？", "補助申請"),
    (48, 4, "幼小銜接前，補助與鑑定安置的時程怎麼排才不會錯過？", "補助申請"),
]

# ============ 新增單輪補題（25 題） ============
NEW_SINGLES = [
    # F 融入日常作息（系統 task F，0 題）
    (49, 1, "孩子放學就累趴了，怎麼把訓練塞進晚間作息又不額外增加負擔？", "融入日常"),
    (50, 2, "吃飯時可以順便練哪些口腔/精細動作？", "融入日常"),
    (51, 3, "洗澡時可以加什麼小遊戲幫助感統發展？", "融入日常"),

    # N 進步查詢（系統 task N，0 題）
    (52, 1, "上次評估到現在三個月了，他有進步嗎？哪些方面進步最多？", "進步查詢"),
    (53, 2, "兩份報告比起來，哪幾個分數有變好、哪幾個沒？", "進步查詢"),
    (54, 3, "他的進步速度算正常嗎？要不要調整訓練？", "進步查詢"),

    # C 臨床觀察補強（原本只有 1 題身心狀態）
    (55, 1, "他常踮腳尖走路，這是感覺尋求還是肌肉張力的問題？", "臨床觀察"),
    (56, 2, "為什麼治療師說他「動作計畫差」？這是什麼意思？", "臨床觀察"),
    (57, 3, "他容易撞到東西、會痛但不哭，是感覺處理異常嗎？", "臨床觀察"),

    # E 居家訓練補強（原本只有 3 題且偏抽象）
    (58, 1, "2 歲半的孩子在家能做哪些精細動作練習？", "居家訓練"),
    (59, 2, "居家訓練每天該做幾次、每次多久才有效又不過度？", "居家訓練"),
    (60, 3, "孩子很抗拒練習，要怎麼設計才能引起他的興趣？", "居家訓練"),

    # 邊界 case：OOD / 閒聊 / 模糊
    (61, 1, "今天天氣不錯耶", "邊界_OOD"),
    (62, 1, "你好，麻煩您了", "邊界_閒聊"),
    (63, 1, "謝謝你的解釋", "邊界_閒聊"),
    (64, 1, "他怎麼了？（無前文脈絡）", "邊界_模糊"),
    (65, 1, "他全方位的發展整體狀況如何？", "邊界_多領域"),
    (66, 1, "哪裡可以做早療？（不講縣市）", "邊界_缺地區"),
    (67, 1, "粗大動作不好會影響認知發展嗎？", "邊界_跨領域"),
    (68, 1, "PR=13 跟標準分 6，意思一樣嗎？", "邊界_術語對照"),

    # SQL 機構/補助實測（剛修台/臺 bug 後重點測）
    (69, 1, "台中哪裡可以做物理治療？", "SQL_機構_半形台"),
    (70, 1, "臺北的早療補助申請流程？", "SQL_補助_全形臺"),
    (71, 1, "新北市有什麼早療資源？", "SQL_機構_市"),
    (72, 1, "台南早療補助一般戶上限多少？", "SQL_補助_額度"),
    (73, 1, "沒有身心障礙證明還能申請什麼補助？", "SQL_補助_無證明"),

    # 報告內容導向補題（對齊 memory_decision_final_combined_1758 語境）
    (74, 1, "報告裡同時寫了能力面向、任務類型和追問情境，我該先看哪個欄位來判斷重點？", "報告內容_欄位解讀"),
    (75, 2, "同一個能力面向下，這次追問和上一輪問題很像，為什麼建議會不同？", "報告內容_追問差異"),
    (76, 3, "如果兩輪問題文字不同但其實在問同一件事，報告該怎麼避免前後矛盾？", "報告內容_一致性"),
    (77, 1, "報告寫這題是『追問型』，家長實際上要怎麼問才比較能得到精準答案？", "報告內容_追問策略"),
    (78, 2, "當問題從粗大動作轉到精細動作，報告會怎麼判斷這是延續還是主題切換？", "報告內容_主題切換"),
    (79, 3, "家長情緒不同（焦慮、困惑、挫折）時，報告內容解讀上會有什麼差異？", "報告內容_情緒語氣"),
    (80, 1, "報告中的『難度』高低，對家長安排訓練優先順序有什麼實際意義？", "報告內容_難度解讀"),
    (81, 2, "同一題如果被標成不同任務類型（例如解釋、建議、流程），我該怎麼理解？", "報告內容_任務類型"),
    (82, 3, "報告裡提到前後文相似度偏低，這代表家長換題了還是系統沒接住？", "報告內容_前後文"),
    (83, 4, "如果報告顯示主題重疊低、距離高，下一輪建議應該怎麼調整問法？", "報告內容_問法調整"),
    (84, 1, "有些題目看起來很短（像『那這樣呢？』），報告是怎麼判斷它在問什麼？", "報告內容_短追問"),
    (85, 2, "我想用報告做家長回饋，哪些欄位最適合拿來追蹤孩子三個月變化？", "報告內容_追蹤應用"),
]

# ============ 多輪 Scenarios（8 個，~33 turns） ============
# scenario_id, turn, query, 面相 / 預期 memory_action 標註
SCENARIOS = [
    # ── S1: Domain 切換鏈（測 v3 連續 REFRESH 能力）
    ("S1", 1, "他現在跑跳正常嗎？", "粗大動作-觀察", "REFRESH"),
    ("S1", 2, "那精細動作呢？", "精細動作-切換", "REFRESH"),
    ("S1", 3, "那認知方面呢？", "認知功能-切換", "REFRESH"),
    ("S1", 4, "三個方面哪個最緊急？需要先處理？", "決策規劃-整合", "REFRESH"),

    # ── S2: 同主題深度追問（測 v3 連續 STAY 能力）
    ("S2", 1, "粗大動作評估結果是什麼？", "粗大動作-評估", "REFRESH"),
    ("S2", 2, "為什麼他單腳站不穩？", "粗大動作-觀察追問", "STAY"),
    ("S2", 3, "那要怎麼訓練平衡？", "粗大動作-訓練", "STAY"),
    ("S2", 4, "具體在家該怎麼做？", "粗大動作-居家", "STAY"),

    # ── S3: 短追問鎖定（測 short_followup_lock 規則）
    ("S3", 1, "他現在的精細動作怎麼樣？", "精細動作-評估", "REFRESH"),
    ("S3", 2, "怎麼訓練？", "短追問", "STAY"),
    ("S3", 3, "為什麼？", "短追問-原因", "STAY"),
    ("S3", 4, "多久才會進步？", "短追問-時程", "STAY"),

    # ── S4: HYBRID — task 切換不換 domain（測 v3 弱點）
    ("S4", 1, "粗大動作評估結果如何？", "粗大動作-評估", "REFRESH"),
    ("S4", 2, "哪裡可以做物理治療？", "task切換H", "REFRESH"),
    ("S4", 3, "大概要排多久？", "task延續H", "STAY"),
    ("S4", 4, "有什麼補助方案？", "task切換K", "REFRESH"),

    # ── S5: 模糊代名詞追問（測 ContextSimilarity）
    ("S5", 1, "報告寫他 PR=13 是什麼意思？", "分數解讀", "REFRESH"),
    ("S5", 2, "那這樣會嚴重嗎？", "代名詞追問", "STAY"),
    ("S5", 3, "為什麼會這樣？", "代名詞追問", "STAY"),
    ("S5", 4, "之後會怎樣？", "代名詞追問", "STAY"),

    # ── S6: 整體 → 單領域 zoom（測 scope 切換）
    ("S6", 1, "報告整體狀況怎麼看？", "整體概況", "REFRESH"),
    ("S6", 2, "粗大動作部分呢？", "scope切換", "REFRESH"),
    ("S6", 3, "他比同齡差幾個月？", "STAY-同領域", "STAY"),
    ("S6", 4, "那精細動作呢？", "domain切換", "REFRESH"),

    # ── S7: 補助→機構連續查詢（測 SQL 連續觸發 + region carry-over）
    ("S7", 1, "台北市的早療補助有哪些？", "補助申請-台北", "REFRESH"),
    ("S7", 2, "申請要備哪些文件？", "STAY-同主題追問", "STAY"),
    ("S7", 3, "那台北哪裡有物理治療所？", "task切換H", "REFRESH"),
    ("S7", 4, "等候時間大約多久？", "STAY-延續", "STAY"),

    # ── S8: 介入評估到家校合作（跨 task 多步推理）
    ("S8", 1, "他現在介入頻率夠嗎？", "介入評估", "REFRESH"),
    ("S8", 2, "應該加課還是換方向？", "介入調整", "STAY"),
    ("S8", 3, "怎麼跟學校老師說明這些調整？", "家校合作", "REFRESH"),
    ("S8", 4, "學校通常能配合什麼？", "STAY-延續", "STAY"),
]

SCENARIO_DESC = {
    "S1": "Domain 切換鏈 — 連續切換 3 個能力領域，最後整合決策。測 v3 連續 REFRESH 判斷力。",
    "S2": "同主題深度追問 — 同一領域從評估→觀察→訓練→居家逐步深入。測 v3 連續 STAY 判斷力。",
    "S3": "短追問鎖定 — 極短代名詞 query。測 short_followup_lock 規則 + ContextSimilarity。",
    "S4": "HYBRID 場景 — 同 domain 但 task 切換到 H/K（機構/補助）。測 v3 已知弱點 + override 救援。",
    "S5": "模糊代名詞追問 — 「那」「這樣」「為什麼」全靠上下文解讀。測語境記憶。",
    "S6": "整體 → 單領域 zoom — scope 從 overview 切換到單 domain 再切其他 domain。",
    "S7": "補助→機構連續查詢 — 跨 task (K↔H) + region 繼承。測 SQL 機構/補助 + 台/臺 normalize。",
    "S8": "介入到家校合作 — 跨 task 多步推理，測 Planning Agent 整合多 section。",
}

# ============ 規範說明（評分標準） ============
RUBRIC = [
    ("資料正確性 (1-5)", "回答中事實、數字、領域名稱是否正確；是否與報告或公認知識相符。",
     "1=多處錯誤誤導 / 3=部分正確 / 5=完全正確"),
    ("連貫性 (1-5)", "段落間邏輯連貫，多輪對話前後語氣一致；語意通順無跳躍。",
     "1=前後矛盾 / 3=尚可閱讀 / 5=結構流暢"),
    ("可行性 (1-5)", "建議是否家長能在家或日常情境執行；步驟具體不空泛。",
     "1=空泛口號 / 3=可執行但需翻譯 / 5=具體可立即做"),
    ("安全性 (1-5)", "建議無誤導風險、無危險動作；對醫療轉介界線清楚。",
     "1=有風險 / 3=邊緣模糊 / 5=完全安全"),
    ("完整性 (1-5)", "回答是否涵蓋家長真正關切；有無漏掉重要面向。",
     "1=嚴重缺漏 / 3=主要點到 / 5=面面俱到"),
    ("整體 (1-5)", "綜合滿意度；如治療師會願意給家長看。",
     "1=不堪用 / 3=可接受 / 5=高品質"),
    ("引用適當性 (1-5)", "資料來源區塊是否合理（領域>段落正確、外部資料有標來源）。",
     "1=亂引或缺引 / 3=部分對 / 5=精準對應"),
    ("長度適當性 (1-5)", "回答篇幅與問題複雜度匹配（不冗長、不過短）。",
     "1=極不匹配 / 3=略長略短 / 5=恰到好處"),
    ("多輪一致性 (1-5)", "[僅多輪 scenario 評] 後輪回答是否合理銜接前輪上下文，不重複也不漏接。",
     "1=完全失憶 / 3=部分接續 / 5=完美延續"),
]

# ============ 寫入 Excel ============

wb = Workbook()

# === Sheet 1: 單輪題庫 ===
ws1 = wb.active
ws1.title = "單輪題庫"

cols_single = [
    ("序號", 6), ("輪次", 6), ("家長問題", 50), ("面相", 14),
    ("資料正確性", 10), ("連貫性", 9), ("可行性", 9), ("安全性", 9),
    ("完整性", 9), ("整體", 8), ("引用適當性", 11), ("長度適當性", 11),
    ("備註", 30),
]
for i, (h, w) in enumerate(cols_single, 1):
    cell = ws1.cell(row=1, column=i, value=h)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = CENTER
    cell.border = BORDER
    ws1.column_dimensions[get_column_letter(i)].width = w
ws1.freeze_panes = "A2"
ws1.row_dimensions[1].height = 28

row = 2
for q in EXISTING + NEW_SINGLES:
    seq, turn, query, aspect = q
    vals = [seq, turn, query, aspect, None, None, None, None, None, None, None, None, None]
    for c, v in enumerate(vals, 1):
        cell = ws1.cell(row=row, column=c, value=v)
        cell.font = CELL_FONT
        cell.border = BORDER
        cell.alignment = WRAP if c == 3 else CENTER
        if c == 4 and aspect.startswith(("邊界", "SQL")):
            cell.fill = ASPECT_FILL
    row += 1

# === Sheet 2: 多輪 Scenarios ===
ws2 = wb.create_sheet("多輪Scenarios")

cols_multi = [
    ("Scenario", 10), ("Turn", 6), ("家長問題", 45), ("面相/Memory標註", 22),
    ("預期Memory", 12),
    ("資料正確性", 10), ("連貫性", 9), ("可行性", 9), ("安全性", 9),
    ("完整性", 9), ("整體", 8), ("引用適當性", 11), ("長度適當性", 11),
    ("多輪一致性", 11), ("備註", 28),
]
for i, (h, w) in enumerate(cols_multi, 1):
    cell = ws2.cell(row=1, column=i, value=h)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = CENTER
    cell.border = BORDER
    ws2.column_dimensions[get_column_letter(i)].width = w
ws2.freeze_panes = "A2"
ws2.row_dimensions[1].height = 28

# Scenario block + description rows
row = 2
last_scenario = None
for s in SCENARIOS:
    sid, turn, query, aspect, expected = s
    if sid != last_scenario:
        # 加 scenario 描述列
        desc = SCENARIO_DESC.get(sid, "")
        ws2.cell(row=row, column=1, value=f"[{sid}] {desc}")
        ws2.cell(row=row, column=1).fill = SCENARIO_FILL
        ws2.cell(row=row, column=1).font = Font(name="Microsoft JhengHei", size=10, bold=True)
        ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=15)
        ws2.cell(row=row, column=1).alignment = WRAP
        row += 1
        last_scenario = sid
    vals = [sid, turn, query, aspect, expected,
            None, None, None, None, None, None, None, None, None, None]
    for c, v in enumerate(vals, 1):
        cell = ws2.cell(row=row, column=c, value=v)
        cell.font = CELL_FONT
        cell.border = BORDER
        cell.alignment = WRAP if c == 3 else CENTER
    row += 1

# === Sheet 3: 評分說明 ===
ws3 = wb.create_sheet("評分說明")

# 標題
ws3.cell(row=1, column=1, value="EduBot v7 專家評分說明")
ws3.cell(row=1, column=1).font = Font(name="Microsoft JhengHei", size=14, bold=True, color="2F5C7E")
ws3.merge_cells("A1:C1")

# 介紹文字
intro_text = (
    "本問卷評估 EduBot v7 早療諮詢系統的回答品質。\n"
    "請就每題回答給予 1-5 分（1=最差，5=最佳）。\n"
    "「單輪題庫」共 73 題（48 既有 + 25 補題）；「多輪Scenarios」8 個對話腳本共 33 turns。\n"
    "每題評 8 維（單輪）或 9 維（多輪含一致性），備註欄可寫具體建議或槽點。\n\n"
    "本次新增評估維度：引用適當性、長度適當性、多輪一致性，請特別留意。"
)
ws3.cell(row=2, column=1, value=intro_text)
ws3.cell(row=2, column=1).alignment = Alignment(wrap_text=True, vertical="top")
ws3.cell(row=2, column=1).font = CELL_FONT
ws3.merge_cells("A2:C2")
ws3.row_dimensions[2].height = 110

# Rubric 表頭
hdr_row = 4
hdrs = ["評估維度", "說明", "評分區間參考"]
widths = [16, 50, 36]
for i, (h, w) in enumerate(zip(hdrs, widths), 1):
    cell = ws3.cell(row=hdr_row, column=i, value=h)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = CENTER
    cell.border = BORDER
    ws3.column_dimensions[get_column_letter(i)].width = w

for i, (dim, desc, scale) in enumerate(RUBRIC, hdr_row + 1):
    ws3.cell(row=i, column=1, value=dim)
    ws3.cell(row=i, column=2, value=desc)
    ws3.cell(row=i, column=3, value=scale)
    for c in range(1, 4):
        ws3.cell(row=i, column=c).font = CELL_FONT
        ws3.cell(row=i, column=c).border = BORDER
        ws3.cell(row=i, column=c).alignment = WRAP
    ws3.row_dimensions[i].height = 42

# 統計摘要區塊
stats_row = hdr_row + len(RUBRIC) + 3
ws3.cell(row=stats_row, column=1, value="題庫覆蓋統計")
ws3.cell(row=stats_row, column=1).font = Font(name="Microsoft JhengHei", size=12, bold=True)

from collections import Counter
all_aspects = [q[3] for q in EXISTING + NEW_SINGLES]
cnt = Counter(all_aspects)
for i, (asp, n) in enumerate(sorted(cnt.items(), key=lambda x: -x[1]), stats_row + 2):
    ws3.cell(row=i, column=1, value=asp).font = CELL_FONT
    ws3.cell(row=i, column=2, value=n).font = CELL_FONT
    ws3.cell(row=i, column=2).alignment = CENTER

# 多輪 scenario 統計
sc_row = stats_row + len(cnt) + 4
ws3.cell(row=sc_row, column=1, value="多輪 Scenario 一覽").font = Font(name="Microsoft JhengHei", size=12, bold=True)
for i, (sid, desc) in enumerate(SCENARIO_DESC.items(), sc_row + 2):
    ws3.cell(row=i, column=1, value=sid).font = Font(name="Microsoft JhengHei", size=10, bold=True)
    ws3.cell(row=i, column=2, value=desc).font = CELL_FONT
    ws3.cell(row=i, column=2).alignment = WRAP
    ws3.row_dimensions[i].height = 30

# 儲存
wb.save(OUT)
print(f"OK: {OUT}")
print(f"  單輪題庫:       {len(EXISTING) + len(NEW_SINGLES)} 題")
print(f"  多輪 Scenarios: {len(SCENARIOS)} turns / {len(SCENARIO_DESC)} scenarios")
print(f"  評分維度:       {len(RUBRIC)} 維")
